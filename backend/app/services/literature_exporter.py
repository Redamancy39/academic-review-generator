# Literature Exporter Service - export papers to various formats
import re
from dataclasses import asdict
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Sequence

from ..core.models import PaperRecord


class LiteratureExporter:
    """Service for exporting literature records to various formats."""

    def __init__(self) -> None:
        """Initialize the exporter."""
        pass

    def export_to_xlsx(
        self,
        records: Sequence[PaperRecord],
        task_id: str = "",
        topic: str = "",
        include_notes: bool = True,
    ) -> bytes:
        """Export records to Excel xlsx format.

        Args:
            records: List of paper records.
            task_id: Task ID for metadata.
            topic: Topic for metadata.
            include_notes: Whether to include a notes column.

        Returns:
            Bytes of the xlsx file.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "文献列表"

        # Define styles
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers
        headers = [
            "序号",
            "标题",
            "作者",
            "年份",
            "期刊",
            "JCR分区",
            "影响因子",
            "被引次数",
            "DOI",
            "摘要",
            "关键词",
            "相关性得分",
            "数据来源",
            "是否入选",
            "备注",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_idx, record in enumerate(records, start=2):
            data = self._record_to_row(record, row_idx - 1)
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Adjust column widths
        column_widths = {
            1: 8,   # 序号
            2: 50,  # 标题
            3: 25,  # 作者
            4: 8,   # 年份
            5: 20,  # 期刊
            6: 10,  # JCR分区
            7: 10,  # 影响因子
            8: 10,  # 被引次数
            9: 25,  # DOI
            10: 60, # 摘要
            11: 25, # 关键词
            12: 12, # 相关性得分
            13: 12, # 数据来源
            14: 10, # 是否入选
            15: 20, # 备注
        }
        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add metadata sheet
        if task_id or topic:
            ws_meta = wb.create_sheet(title="导出信息")
            ws_meta["A1"] = "导出时间"
            ws_meta["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws_meta["A2"] = "任务ID"
            ws_meta["B2"] = task_id
            ws_meta["A3"] = "综述主题"
            ws_meta["B3"] = topic
            ws_meta["A4"] = "文献总数"
            ws_meta["B4"] = len(records)
            ws_meta.column_dimensions["A"].width = 15
            ws_meta.column_dimensions["B"].width = 50

        # Write to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def export_to_csv(
        self,
        records: Sequence[PaperRecord],
        task_id: str = "",
        topic: str = "",
    ) -> str:
        """Export records to CSV format.

        Args:
            records: List of paper records.
            task_id: Task ID for metadata.
            topic: Topic for metadata.

        Returns:
            CSV string.
        """
        import csv
        from io import StringIO

        output = StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL)

        # Headers
        headers = [
            "序号", "标题", "作者", "年份", "期刊", "JCR分区",
            "被引次数", "DOI", "摘要", "关键词", "相关性得分", "数据来源"
        ]
        writer.writerow(headers)

        # Data
        for idx, record in enumerate(records, start=1):
            writer.writerow(self._record_to_row(record, idx)[:12])

        return output.getvalue()

    def export_to_bibtex(
        self,
        records: Sequence[PaperRecord],
    ) -> str:
        """Export records to BibTeX format.

        Args:
            records: List of paper records.

        Returns:
            BibTeX string.
        """
        lines = []
        for record in records:
            cite_key = self._generate_cite_key(record)
            authors = " and ".join(record.authors[:5]) if record.authors else "Unknown"
            year = record.year or "n.d."

            lines.append(f"@article{{{cite_key},")
            lines.append(f"  author = {{{authors}}},")
            lines.append(f"  title = {{{record.title}}},")
            lines.append(f"  journal = {{{record.journal}}},")
            lines.append(f"  year = {{{year}}},")
            if record.doi:
                lines.append(f"  doi = {{{record.doi}}},")
            if record.url:
                lines.append(f"  url = {{{record.url}}},")
            lines.append("}")
            lines.append("")

        return "\n".join(lines)

    def _record_to_row(self, record: PaperRecord, index: int) -> List[Any]:
        """Convert a record to a row for Excel/CSV export.

        Args:
            record: Paper record.
            index: Row index.

        Returns:
            List of values for the row.
        """
        return [
            f"REF{index:03d}",  # 序号
            record.title,  # 标题
            ", ".join(record.authors[:5]) if record.authors else "",  # 作者
            record.year or "",  # 年份
            record.journal,  # 期刊
            record.jcr_quartile,  # JCR分区
            getattr(record, "impact_factor", ""),  # 影响因子 (P2阶段添加)
            record.times_cited,  # 被引次数
            record.doi,  # DOI
            record.abstract[:500] if record.abstract else "",  # 摘要
            ", ".join(record.keywords[:10]) if record.keywords else "",  # 关键词
            round(record.relevance_score, 2),  # 相关性得分
            record.source_db,  # 数据来源
            "",  # 是否入选 (用户填写)
            "",  # 备注 (用户填写)
        ]

    def _generate_cite_key(self, record: PaperRecord) -> str:
        """Generate a citation key for BibTeX.

        Args:
            record: Paper record.

        Returns:
            Citation key string.
        """
        # Use first author's last name + year + first word of title
        if record.authors:
            first_author = record.authors[0].split()[-1].lower()
            first_author = re.sub(r"[^a-z]", "", first_author)
        else:
            first_author = "unknown"

        year = record.year or "xxxx"

        # First significant word from title
        title_words = re.findall(r"[A-Za-z]+", record.title)
        first_word = title_words[0].lower() if title_words else "paper"

        return f"{first_author}{year}{first_word}"
